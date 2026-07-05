from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook


def cell_value_to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_excel(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    items: list[dict] = []
    for sheet in workbook.worksheets:
        rows = []
        min_row = None
        max_row = None
        min_col = None
        max_col = None
        for row in sheet.iter_rows():
            values = [cell_value_to_text(cell.value) for cell in row]
            if not any(values):
                continue
            if min_row is None:
                min_row = row[0].row
            max_row = row[0].row
            used_cols = [cell.column for cell in row if cell_value_to_text(cell.value)]
            if used_cols:
                min_col = min(used_cols) if min_col is None else min(min_col, min(used_cols))
                max_col = max(used_cols) if max_col is None else max(max_col, max(used_cols))
            rows.append(" | ".join(values))
        if rows:
            cell_range = None
            if min_row and max_row and min_col and max_col:
                cell_range = f"R{min_row}C{min_col}:R{max_row}C{max_col}"
            items.append({
                "text": "\n".join(rows),
                "section": f"Sheet: {sheet.title}",
                "sheet_name": sheet.title,
                "cell_range": cell_range,
                "extraction_method": "openpyxl",
            })
    return items
